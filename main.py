import openpyxl
import datetime

wb2 = openpyxl.load_workbook('C:\\Users\\Doug Trefun\\Desktop\\call_stats\\phone_data.xlsx')
sheet = wb2.active
max_col = sheet.max_column
max_row = sheet.max_row


#columns from xlsx file into list
col_lst = []
for i in range(1,max_col+1):
    cell_obj = sheet.cell(row=1,column=i)
    col_lst.append(cell_obj.value)


#rows from xlsx file into list
row_lst = []
for j in range(1,max_col+1):
    cell_obj2 = sheet.cell(row=3, column= j)
    row_lst.append(cell_obj2.value)

print(type(row_lst))

#converting time to minutes
time_string = ""
times_fixed = []
for x in row_lst:
    if type(x) == datetime.time:
        time_string = str(x)
        date_time = datetime.datetime.strptime(time_string, "%H:%M:%S")
        a_timedelta = date_time - datetime.datetime(1900,1,1)
        seconds = a_timedelta.total_seconds()
        times_fixed.append(round(seconds/60,2))
    else:
        times_fixed.append(x)


#combining lists into dict
result_dict = dict(zip(col_lst,times_fixed))

#iterating through dict

for key,value in result_dict.items():
    print(key, ': ', value)

time_items = []
for x in result_dict.values():
    if type(x) == float:
        time_items.append(x)
print((sum(time_items)/60))


total_total = sum(time_items)/60 - max(time_items)/60


